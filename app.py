import streamlit as st
import pandas as pd
from openpyxl import Workbook
import matplotlib.pyplot as plt
from io import BytesIO

# Title and instruction
st.title('Breakeven Calculator')
st.markdown("Download the template, fill it with your data (Year in column A, Cash Flow in column B, Initial Investment in cell C2), and upload it.")

# Function to create an Excel template
def create_excel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Data"
    ws['A1'] = 'Year'
    ws['B1'] = 'Cash Flow'
    ws['C1'] = 'Initial Investment'
    data = [(1, 2000), (2, 3000), (3, 3500), (4, 4000)]
    for i, (year, cash_flow) in enumerate(data, start=2):
        ws[f'A{i}'] = year
        ws[f'B{i}'] = cash_flow
    ws['C2'] = 10000
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

# Download Excel template
excel_template = create_excel_template()
st.download_button(
    label="Download Excel Template",
    data=excel_template,
    file_name="Financial_Data_Template.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)

# Upload filled Excel file
uploaded_file = st.file_uploader("Upload your filled Excel file here:", type=['xlsx'])

if uploaded_file:
    # Load the Excel file into a DataFrame
    df = pd.read_excel(uploaded_file, engine='openpyxl')

    # Display data editor
    edited_df = st.data_editor(df, use_container_width=True)

    # After editing, plot and display the new data
    if st.button('Update Chart'):

        st.divider()

        initial_investment = edited_df['Initial Investment'][0]  # Assumes the first row contains initial investment
        years = edited_df['Year'].tolist()
        cash_flows = edited_df['Cash Flow'].tolist()

        # Calculating cumulative cash flow
        cumulative_cash_flow = []
        cumulative = 0
        for cash in cash_flows:
            cumulative += cash
            cumulative_cash_flow.append(cumulative)

        # Find the break-even point with decimal precision
        break_even_year = None
        for i in range(1, len(cumulative_cash_flow)):
            if cumulative_cash_flow[i-1] < initial_investment <= cumulative_cash_flow[i]:
                # Linear interpolation
                break_even_year = years[i-1] + (initial_investment - cumulative_cash_flow[i-1]) / (cumulative_cash_flow[i] - cumulative_cash_flow[i-1])
                break


        # Display the break-even point
        if break_even_year:
            st.metric(label="Break-Even Year", value=f"Year {break_even_year}", delta=None, delta_color="inverse")
        else:
            st.metric(label="Break-Even Year", value="No break-even within the given years", delta=None, delta_color="off")

        # Creating the plot
        plt.figure(figsize=(10, 6))
        plt.plot(years, cumulative_cash_flow, marker='o', label='Cumulative Cash Flow')
        plt.axhline(y=initial_investment, color='r', linestyle='--', label='Initial Investment')
        plt.title('Investment Payback Period')
        plt.xlabel('Year')
        plt.ylabel('Cumulative Cash Flow')
        plt.legend()
        plt.grid(True)
        st.pyplot(plt)
